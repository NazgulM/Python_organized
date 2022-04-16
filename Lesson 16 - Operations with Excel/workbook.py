import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('tom.xlsx')
ws = wb.active
# ws.append(['Tom', 'Is', 'Great', '!'])
# ws.append(['Tom', 'Is', 'Great', '!'])
# ws.append(['Tom', 'Is', 'Great', '!'])
# ws.append(['Tom', 'Is', 'Great', '!'])
# ws.append(['end'])
# for row in range(1, 11):
#     for col in range(1, 5):
#         char = get_column_letter(col)
#         # print(ws[char + str(row)])  # char = chr(65 + col)
#         ws[char + str(row)] = char + str(row)

# ws.merge_cells('A1:D1')
# ws.unmerge_cells('A1:D1')
# ws.merge_cells('A1:D2')
# ws.insert_rows(7)
# ws.insert_rows(7)
# ws.delete_rows(7)
# ws.insert_cols(2)
# ws.delete_cols(2)
ws.move_range('C1:D11', rows=2, cols=2)
wb.save('tom.xlsx')
