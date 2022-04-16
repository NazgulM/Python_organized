import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import *

wb = load_workbook('../Lesson 18 - Repeat File Excel/food.xlsx')
sheet = wb.active
print("Успешно!")
# Прочтите файл excel
# print(sheet['A1'].value)
# print()
for column in sheet['A']:
    print(column.value)
#  Из файла прочтите «первую колонку» данных
# cells = sheet['A1': 'A245']
# print(cells)

dates = sheet['A']
print(dates)
dateList = list()
for date in dates:
    if date.value == 'OrderDate':
        continue
    dateList.append(date.value)
print(dateList)
print()

# Получите значение F19
print(sheet['F19'].value)
print()

# Получите значение в диапазоне [‘A18:E30’] и назначьте ее в
# значение myListExcel
cells = sheet['A18': 'H18']
# cellsData = cells[0]
print(cells)
# print(cellsData)
print(len(cells[0]))
for info1, info2, info3, info4, info5, info6, info7, info8 in cells:
    print(info1.value, info2.value, info3.value, info4.value, info5.value, info6.value, info7.value, info8.value)

# 7. Сделайте перебор данных установив лимит в 5 рядов из myListExcel
# print((cells[:6]))
# list1 = list(cells[:6])
# print(list1)
# print(type(list1))
# print(len(list1[0]))
# print(list1[:6])
for data in list(cells[0][:6]):
    print(data.value, end=' ')

myTuple = ((2, 32, 4),)
print(list(myTuple[0]))
'''

8. Попросить вести пользователя следующие данные, как в след. примере:
[‘Имя’,’Возраст’,’Зарплата’,’Год работ’], 
[‘Самат’, ‘Улан’]
[25, 30]
[25000, 32000]
[5,7]
'''
wb1 = Workbook()
sheet1 = wb1.worksheets[0]
print()
menu = []
n = int(input('Please enter the number of elements: '))

for i in range(0, n):
    elem = input('Please enter: ')
    menu.append(elem)
print(menu)
print()

names = []
n1 = int(input('Please enter the number of elements: '))

for j in range(0, n1):
    elem1 = input('Please enter names: ')
    names.append(elem1)
print(names)
print()

ages = []
n2 = int(input('Please enter the number of elements: '))

for k in range(0, n2):
    elem2 = input('Please enter ages: ')
    ages.append(elem2)
print(ages)
print()

salary = []
n3 = int(input('Please enter the number of elements: '))

for a in range(0, n3):
    elem3 = input('Please enter salaries: ')
    salary.append(elem3)
print(salary)
print()

exp = []
n4 = int(input('Please enter the number of elements: '))

for a in range(0, n4):
    elem4 = input('Please enter employees exp.years: ')
    exp.append(elem4)
print(exp)

# column1Name = sheet.cell(row=1, column=1).value = 'Name of Employee'
# column2Name = sheet.cell(row=1, column=2).value = 'Age'
# column3Name = sheet.cell(row=1, column=3).value = 'Salary'
# column4Name = sheet.cell(row=1, column=4).value = 'Experience'

count = 1
for j in range(4):
    sheet1.cell(row=1, column=count).value = menu[j]
    sheet1.cell(row=1, column=count).value = menu[j]
    sheet1.cell(row=1, column=count).value = menu[j]
    sheet1.cell(row=1, column=count).value = menu[j]
    count +=1

counter = 2
for i in range(2):
    sheet1.cell(row=counter, column=1).value = names[i]
    sheet1.cell(row=counter, column=2).value = ages[i]
    sheet1.cell(row=counter, column=3).value = salary[i]
    sheet1.cell(row=counter, column=4).value = exp[i]
    counter += 1

wb1.save('workers.xlsx')
print('Successfully saved our file')
