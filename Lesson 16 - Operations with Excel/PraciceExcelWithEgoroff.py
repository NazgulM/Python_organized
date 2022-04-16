import openpyxl
from openpyxl import load_workbook

book = openpyxl.open('../Lesson 18 - Repeat File Excel/food.xlsx', read_only=True)
sheet = book.active

print(sheet['B2'].value)
print(sheet[1][0].value)

cells = sheet['B1': 'C11']
for region, city in cells:
    print(region.value, city.value)

# for row in range(1, sheet.max_row+1):
#     date = sheet[row][0].value
#     region = sheet[row][1].value
#     city = sheet[row][2].value
#     category = sheet[row][3].value
#     product = sheet[row][4].value
#     quantity = sheet[row][5].value
#     unit = sheet[row][6].value
#     totalPrice = sheet[row][7].value
#     print(date, region, city, category, product, quantity, unit,totalPrice)


wb = load_workbook('../Lesson 18 - Repeat File Excel/food.xlsx')
ws = wb.active
# print(ws)
# print()
# print(ws['A1'].value)
# ws['A1'].value = 'Order Date'
# wb.save('food.xlsx')
# print()
#
# # print(wb.sheetnames)
# # # wb.create_sheet('Test')
# # # print(wb.sheetnames)
# # # wb.save('food.xlsx')
