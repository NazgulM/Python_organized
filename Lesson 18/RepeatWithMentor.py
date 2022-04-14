from openpyxl import *
"""
Task 2.
Отобразите все уникальные название городов, категорий, регионов из файла ‘food.xlsx’
"""
def showUnique(sheet):

    #unique regions
    regions = sheet['B']
    regionList = [region.value for region in regions if region.value != 'Region']
    # for region in regions:
    #     regionList.append(region.value)
    regionList = list(set(regionList))
    print(f'Unique regions: {regionList}')

    # unique cities
    cities = sheet['C']
    citiesList = [city.value for city in cities if city.value != 'City']
    citiesList = list(set(citiesList))
    print(f'Unique cities: {citiesList}')

    # unique categories
    categories = sheet['D']
    categoriesList = [category.value for category in categories if category.value != 'Category']
    categoriesList = list(set(categoriesList))
    print(f'Unique categories: {categoriesList}')

def schoolFile(filename):
    wb2 = Workbook()
    wb2.create_sheet('marks')
    sheet1 = wb2['Sheet']
    wb2.remove(sheet1)
    sheet = wb2['marks']

    headerList = ['Name', 'Mathematic', 'Chemistry', 'Geography', 'Algebra', 'Average', 'Sum']

    quantStudent = int(input('How many students you want to input?: '))

    dataList = []

    for i in range(quantStudent):
        name = input('Put his name: ')
        math = int(input('Mark for Math:'))
        chemistry = int(input('Mark for Chemistry:'))
        geography = int(input('Mark for Geography:'))
        algebra = int(input('Mark for Algebra:'))

        dataList.append([name, math, chemistry, geography, algebra])

    dataList.insert(0, headerList)

    for data in dataList:
        sheet.append(data)

    counter = 2
    for data in range(quantStudent):
        sheet[f'G{counter}'] = f'= sum(B{counter}:E{counter})'
        sheet[f'F{counter}'] = f'= AVERAGE(B{counter}:E{counter})'
        counter += 1

    wb2.save(filename)

def main():
    filename = 'food.xlsx'
    filename2 = 'school2.xlsx'
    wb = load_workbook(filename)
    sheet = wb.worksheets[0]
    showUnique(sheet)

    schoolFile(filename2)


if __name__ == '__main__':
    main()