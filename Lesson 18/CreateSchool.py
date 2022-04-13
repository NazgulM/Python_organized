import openpyxl
from openpyxl import *

'''
Создайте excel файл ‘school.xlsx’. В этом файле:
a) Создайте страницу ‘marks’
b) Создайте следующие заголовки:

• Имя школьника
• Математика
• География
• Астрономия
• Химия
• Физика
'''

# def newFile():
#     wb = Workbook()
#     sheet = wb.active
#     dataMarks = [('Name of student', 'Math', 'Geography', 'Astronomy', 'Chemistry', 'Physics')]
#     for row in dataMarks:
#         sheet.append(row)
#     wb.save('school.xlsx')
#
#     return newFile()
#
#
# def main():
#    pass
#
#
# if __name__ == '__main__':
#     main()

# wb = Workbook()
# sheet = wb.active
# dataMarks = [('Name of student', 'Math', 'Geography', 'Astronomy',
#               'Chemistry', 'Physics')]
# for row in dataMarks:
#     sheet.append(row)


'''
c) Примите от пользователя количество учеников, которых необходимо добавить в 
данный файл, путем цикла примите данные 
для этих учеников и заполните исходя из заголовка задания b)
'''
# names = []
# num = int(input('Please enter the number of elements: '))
#
# for i in range(0, num):
#     elem = input('Please enter the names of students: ')
#     names.append(elem)
# print(names)
#
# mathMarks = []
# num = int(input('Please enter the number of elements: '))
#
# for i in range(0, num):
#     elem = int(input('Please enter the marks for mathematics: '))
#     mathMarks.append(elem)
# print(mathMarks)
#
# geography = []
# num = int(input('Please enter the number of elements: '))
#
# for i in range(0, num):
#     elem = int(input('Please enter the marks for geography: '))
#     geography.append(elem)
# print(geography)
#
# astronomy = []
# num = int(input('Please enter the number of elements: '))
#
# for i in range(0, num):
#     elem = int(input('Please enter the marks for astronomy: '))
#     astronomy.append(elem)
# print(astronomy)
#
# chemistry = []
# num = int(input('Please enter the number of elements: '))
#
# for i in range(0, num):
#     elem = int(input('Please enter the marks for chemistry: '))
#     chemistry.append(elem)
# print(chemistry)
#
# physics = []
# num = int(input('Please enter the number of elements: '))
#
# for i in range(0, num):
#     elem = int(input('Please enter the marks for physics: '))
#     physics.append(elem)
# print(physics)
#
# counter = 2
# for i in range(5):
#     sheet.cell(row=counter, column=1).value = names[i]
#     sheet.cell(row=counter, column=2).value = mathMarks[i]
#     sheet.cell(row=counter, column=3).value = geography[i]
#     sheet.cell(row=counter, column=4).value = astronomy[i]
#     sheet.cell(row=counter, column=5).value = chemistry[i]
#     sheet.cell(row=counter, column=6).value = physics[i]
#
#     counter += 1
#
# wb.save('school.xlsx')

'''
d) Используя формулу запишите в отдельные поля:
• Итоговая оценка (Всех учеников)
• Средняя оценка (Для всех учеников)
'''

book = openpyxl.load_workbook('school.xlsx')
sheet = book.active
sheet['G1'] = 'Total grades'
sheet['G2'] = '= sum(B2:F2)'
sheet['G3'] = '= sum(B3:F3)'
sheet['G4'] = '= sum(B4:F4)'
sheet['G5'] = '= sum(B5:F6)'
sheet['G6'] = '= sum(B6:F6)'

sheet['H1'] = 'Average grade'
sheet['H2'] = '= AVERAGE(B2:F2)'
sheet['H3'] = '= AVERAGE(B3:F3)'
sheet['H4'] = '= AVERAGE(B4:F4)'
sheet['H5'] = '= AVERAGE(B5:F5)'
sheet['H6'] = '= AVERAGE(B6:F6)'


book.save('school.xlsx')