from openpyxl import *

"""
1. ['F19':'G30'] - Прочитать в список данный промежуток из ‘food.xlsx’
Найти:
• Минимальное значение
• Максимальное значение
• Среднеарифметическое значение
"""


def minValue():
    wb = load_workbook('food.xlsx')
    sheet = wb.worksheets[0]
    sheetName = wb['Food_List']
    cells = sheetName['F19':'G30']

    values = []

    for row in cells:
        for data in row:
            values.append(data.value)

    return values


def maxValue():
    wb1 = load_workbook('food.xlsx')
    sheet1 = wb1.worksheets[0]
    sheetName = wb1['Food_List']
    cells = sheetName['F19':'G30']

    values1 = []

    for row in cells:
        for data in row:
            values1.append(data.value)

    return values1


def average():
    wb2 = load_workbook('food.xlsx')
    sheet2 = wb2.worksheets[0]
    sheetName = wb2['Food_List']
    cells2 = sheetName['F19':'G30']

    values2 = []

    for row in cells2:
        for data in row:
            values2.append(data.value)
        sumList = sum(values2)
        averageList = sumList / len(values2)

    return averageList


'''
2. Отобразите все уникальные название городов, категорий, регионов из файла ‘food.xlsx’
'''


def uniqueNames():
    wb3 = load_workbook('food.xlsx')
    sheet3 = wb3.worksheets[0]
    sheetName = wb3['Food_List']
    cells3 = sheetName['B2':'D245']

    values3 = []

    for row in cells3:
        for data in row:
            values3.append(data.value)
        set1 = set(values3)

    return set1



def main():
    # Task 1
    print(f'The minimum number is: {min(minValue())}')

    # Task 2
    print(f'The maximum number is: {max(maxValue())}')

    # Task 3
    print(f'The average of numbers: {average()}')

    # Task 4
    print(uniqueNames())




if __name__ == '__main__':
    main()
