import stats as stats
from openpyxl import *
import statistics as stat

'''
1) Read range using 2 options:
- by coordinates
- by row and columns ['B12':'G28']
2) ['F19':'G30'] - read in list
get 
max, 
median, 
mean
variance
deviation
3) Write data to Excel
Info about Kyrgyzstan
 - Name of Regions
 - Name of big cities of regions
  - Population in thousands(100, 200)
  - Mayor of the city
'''

wb = load_workbook('food.xlsx')
sheet = wb.worksheets[0]

sheetName = wb['Food_List']
cells = sheetName['B12':'G28']

for row in cells:
    for data in row:
        print(data.value, end=" ")
    print(end='\n')

print('*' * 20)

cells2 = sheetName['B12':'G28']
cells2 = sheetName.iter_rows(min_row=12, max_col=7, max_row=28)
listExcel = []
for row in cells2:
    for data in row:
        listExcel.append(data.value)
    # print(listExcel, end=' ')
    # print(end='\n')
listExcel2 = listExcel[1:]
print(listExcel2, end=' ')
print(end='\n')

print()

sheetName = wb['Food_List']
cells2 = sheetName['F19':'G30']
values = []

for row in cells2:
    for data in row:
        values.append(data.value)

print(values)
print(f'Length of values: {len(values)}')
print(f'Number of values: {max(values)}')
print(f'Median {stats.median(values)}')
print(f'Variance {stats.variance(values)}')
print(f'St Deviation {stats.stdev(values)}')
print(f'Mean {stats.mean(values)}')
print(f'Mean {sum(values) / len(values)}')

'''
Write data to Excel
Info about Kyrgyzstan
 - Name of Regions
 - Name of big cities of regions
  - Population in thousands(100, 200)
  - Mayor of the city
'''
wb2 = Workbook()
sheet2 = wb2.active
dataCountry= (
    ('Regions of Kyrgyzstan', 'Big city in region', 'Population(thousands)', 'Gubernator'),
    ('Naryn', 'Naryn', '292.14', 'Omurbek Suvanaliev'),
    ('Chuy', 'Bishkek', '1000', 'Abdrahmanov Sagynbek'),
    ('Issyk-Kul', 'Cholpon-Ata', '502', 'Mirbek Asanakunov'),
    ('Jalal-Abad', 'Jalal-Abad', '1300', 'Zhusupbek Sharipov'),
    ('Osh', 'Osh', '1400', 'Taalaybek Sarybashev'),
    ('Batken', 'Batken', '549', 'Mamat Aibalayev'),
    ('Talas', 'Talas', '271', 'Koisun Kurmanalieva')
)

for row in dataCountry:
    sheet2.append(row)

wb2.save('infoKyrgyzstan.xlsx')