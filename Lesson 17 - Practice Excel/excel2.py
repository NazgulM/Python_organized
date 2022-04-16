import stats as stats
from openpyxl import *
import statistics as stat

wb = load_workbook('food.xlsx')
sheet = wb.worksheets[1]
print(sheet)

allSheetNames = wb.sheetnames

print(allSheetNames)

for i in allSheetNames:
    print(i, end=" ")

# display tabs by name
print()
sheetByName = wb['Food_addresses']
print(sheetByName['B2'].value)

# reading big range in Excel
'''
[
[23, 43, 54, 656], - row
[3,45, 65, 6, 76],
[34, 54, 65, 89]
]

for row in someList:
for data in row:
print(data)

'''
print()
sheetName = wb['Food_List']
cells = sheetName['A2':'D17']

for row in cells:
    for data in row:
        print(data.value, end=" ")
    print(end='\n')

# Second option for reading
print('*' * 10)
cells2 = sheetName.iter_rows(min_row=1, max_col=3, max_row=1)
for row in cells2:
    for data in row:
        print(data.value, end=" ")
    print(end='\n')

print('*' * 15)
cells3 = sheetName.iter_rows(min_row=2, max_col=5, max_row=20)
for row in cells3:
    for data in row:
        print(data.value, end=" ")
    print(end='\n')

# reading row by row
print('*' * 20)
for row in sheetName.rows:
    for data in row:
        print(data.value, end=" ")
    print(end="\n")

print()
allRows = sheetName.rows
print(next(allRows))

print()
cells4 = sheetName['E5':'G5']
for row in cells4:
    for data in row:
        print(type(data.value))

# Write multiple data in one time to excel
print()
# wb2 = Workbook()
# sheet2 = wb2.active
# dataExcel = (
#     ('Country', 'Capital', 'Language'),
#     ('Japan', 'Tokyo', 'Japanese'),
#     ('Kyrgyzstan', 'Bishkek', 'Kyrgyz'),
#     ('India', 'Delhi', 'Hindu'),
#     ('Saud Arabia', 'Er Riyad', 'Arabian'),
#     ('USA', 'Washington', 'English'),
# )
#
# for row in dataExcel:
#     sheet2.append(row)
#
# wb2.save('countryInfo.xlsx')

wb = load_workbook('food.xlsx')
sheet = wb.worksheets[3]

allRows = sheet.rows
values = []

for row in allRows:
    for data in row:
        values.append(data.value)

print(values)
print(f'Number of values: {len(values)}')
print(f'Number of values: {max(values[3:])}')
print(values[3:])
listNum = values[3:]
listNum.sort()
print(listNum)
print(f'Median {stats.median(values[3:])}')
print(f'Variance {stats.variance(values[3:])}')
print(f'St Deviation {stats.stdev(values[3:])}')
print(f'Mean {stats.mean(values[3:])}')
print(f'Mean {sum(values[3:])/len(values[3:])}')
