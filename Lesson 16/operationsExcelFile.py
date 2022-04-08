import openpyxl
# from openpyxl import *
from openpyxl import load_workbook, Workbook
from openpyxl.styles import *
wb = load_workbook('example.xlsx')
sheet = wb.active

# reading cell
print(sheet['A2'].value)
name1Person = sheet['A2'].value
age1Person = sheet['B2'].value
print(f'Name of first person: {name1Person} and age of first person: {age1Person}')
print(type(name1Person))
print(type(age1Person))

# reading multiple cells
print()
cells = sheet['A3': 'C3']
print(cells)

for info1, info2, info3 in cells:
    print(info1.value, info2.value, info3.value)

# reading specific column
print()
salaryList = [sheet.cell(row=4, column=3).value, sheet.cell(row=5, column=3).value,
              sheet.cell(row=6, column=3).value]
print(salaryList)

for salary in salaryList:
    print(salary, end=" ")
print()
names = sheet['A']
print(names)

columnNameList = list()
for name in names:
    if name.value == 'Имя':
        continue
    columnNameList.append(name.value)
print(columnNameList)

print()
ages = sheet['B']
print(ages)
columnAgeList = list()
for age in ages:
    if age.value == 'Возраст':
        continue
    columnAgeList.append(age.value)
print(columnAgeList)

listAgeName = [info for info in zip(columnNameList, columnAgeList)]
print(listAgeName)
columnAgeList2 = [age.value for age in ages if age.value == 'Возраст']
print(columnAgeList2)

# read info from Excel
wb2 = Workbook()
sheet1 = wb2.worksheets[0]

nameStudentList = ['Nazgul', 'Kymbat', 'Altynai', 'Nazima', 'Aiperi', 'Ayar']
countryList = ['Belgium', 'Kazakhstan', 'Kyrgyzstan', 'USA', 'South America', 'Russia']
infoStudentEnglishLevel = [80, 82, 75, 67, 78, 100]

sheet1['A1'].font = Font(bold=True)
sheet1['B1'].font = Font(bold=True)
sheet1['C1'].font = Font(bold=True)

column1Name = sheet1.cell(row=1, column=1).value = 'Name of Student'
column2Name = sheet1.cell(row=1, column=2).value = 'Country'
column3Name = sheet1.cell(row=1, column=3).value = 'English level'

counter = 2
for i in range(6):
    sheet1.cell(row=counter, column=1).value = nameStudentList[i]
    sheet1.cell(row=counter, column=2).value = countryList[i]
    sheet1.cell(row=counter, column=3).value = infoStudentEnglishLevel[i]
    counter+=1

wb2.save('infoStudent.xlsx')
print('Successfully saved our file')


